import csv
import io
from flask import Blueprint, request, current_app, render_template, session, send_file
from bson import ObjectId
from app.services.grade_service import GradeService
from app.utils.auth import teacher_required, login_required
from app.utils.helpers import success_response, error_response

grades_bp = Blueprint("grades", __name__)


def get_service():
    return GradeService(current_app.db)


@grades_bp.route("/entry", methods=["GET"])
@teacher_required
def grade_entry():
    user = session.get("user", {})
    uid  = ObjectId(user["id"])

    teacher_doc = current_app.db.teachers.find_one({"user_id": uid})
    if not teacher_doc:
        return render_template("grades/entry.html", courses=[], selected_course=None,
                               student_grades=[], semester="Guz Donemi", academic_year="2024-2025")

    courses = list(current_app.db.courses.find({"teacher_id": teacher_doc["_id"]}))

    course_id = request.args.get("course_id")
    selected_course = None
    student_grades  = []

    if course_id:
        selected_course = current_app.db.courses.find_one({"_id": ObjectId(course_id)})
        if selected_course:
            class_name = selected_course.get("class_name", "")
            students   = list(current_app.db.students.find({"class_name": class_name}).sort("full_name", 1))
            for student in students:
                grade = current_app.db.grades.find_one({
                    "student_id": student["_id"],
                    "course_id": selected_course["_id"],
                })
                student_grades.append({"student": student, "grade": grade})

    return render_template(
        "grades/entry.html",
        courses=courses,
        selected_course=selected_course,
        student_grades=student_grades,
        semester="Guz Donemi",
        academic_year="2024-2025",
        user=user,
    )


@grades_bp.route("/student/<student_id>", methods=["GET"])
@login_required
def student_grades(student_id: str):
    academic_year = request.args.get("academic_year")
    if request.is_json:
        grades = get_service().get_student_grades(student_id, academic_year)
        return success_response(grades)

    # Web görünümü için course bilgisini de join et
    from bson import ObjectId
    from app.utils.helpers import serialize_doc
    query = {"student_id": ObjectId(student_id)}
    if academic_year:
        query["academic_year"] = academic_year
    pipeline = [
        {"$match": query},
        {"$lookup": {"from": "courses", "localField": "course_id",
                     "foreignField": "_id", "as": "course"}},
        {"$unwind": {"path": "$course", "preserveNullAndEmptyArrays": True}},
        {"$sort": {"course.course_name": 1}},
    ]
    grades = serialize_doc(list(current_app.db.grades.aggregate(pipeline)))
    return render_template("grades/student_grades.html", grades=grades, student_id=student_id)


@grades_bp.route("/course/<course_id>", methods=["GET"])
@teacher_required
def course_grades(course_id: str):
    grades = get_service().get_course_grades(course_id)
    return success_response(grades)


@grades_bp.route("/", methods=["POST"])
@teacher_required
def upsert_grade():
    data = request.get_json() or request.form.to_dict()
    result = get_service().upsert_grade(data)
    if not result["success"]:
        return error_response("Not kaydedilemedi.", 400, result.get("errors"))
    status = 201 if result["action"] == "created" else 200
    return success_response({"grade_id": result["grade_id"]}, "Not kaydedildi.", status)


@grades_bp.route("/class/<class_name>/averages", methods=["GET"])
@teacher_required
def class_averages(class_name: str):
    academic_year = request.args.get("academic_year")
    result = get_service().get_class_averages(class_name, academic_year)
    return success_response(result)


@grades_bp.route("/class/<class_name>/ranking", methods=["GET"])
@login_required
def class_ranking(class_name: str):
    academic_year = request.args.get("academic_year")
    result = get_service().get_class_ranking(class_name, academic_year)
    if request.is_json:
        return success_response(result)
    return render_template("grades/ranking.html", ranking=result, class_name=class_name)


@grades_bp.route("/upload-csv", methods=["POST"])
@teacher_required
def upload_csv():
    """
    CSV formatında toplu not yükleme.
    Beklenen sütunlar: ogrenci_no, ders_kodu, vize, proje (opsiyonel), final
    """
    file = request.files.get("csv_file")
    if not file or not file.filename.endswith(".csv"):
        return error_response("Geçerli bir CSV dosyası yükleyin.", 400)

    service = get_service()
    content = file.read().decode("utf-8-sig")  # BOM destekli
    reader  = csv.DictReader(io.StringIO(content))

    # Sütun adı eşleştirme (Türkçe veya İngilizce başlık)
    ALIAS = {
        "ogrenci_no":    ["ogrenci_no","student_number","öğrenci_no","no"],
        "ders_kodu":     ["ders_kodu","course_code","ders"],
        "vize":          ["vize","midterm","midterm_score","vize_notu"],
        "proje":         ["proje","project","project_score","proje_notu"],
        "final":         ["final","final_score","final_notu"],
    }

    def find_col(row, key):
        for alias in ALIAS[key]:
            if alias in row:
                return row[alias]
        return None

    def safe_float(v):
        try:
            f = float(str(v).replace(",", ".").strip())
            return f if 0 <= f <= 100 else None
        except (TypeError, ValueError):
            return None

    success_count = 0
    error_count   = 0
    errors        = []

    for line_no, row in enumerate(reader, start=2):
        student_no  = find_col(row, "ogrenci_no")
        course_code = find_col(row, "ders_kodu")
        midterm     = safe_float(find_col(row, "vize"))
        project     = safe_float(find_col(row, "proje"))
        final       = safe_float(find_col(row, "final"))

        if not student_no or not course_code:
            errors.append(f"Satır {line_no}: öğrenci no veya ders kodu eksik.")
            error_count += 1
            continue

        student = current_app.db.students.find_one({"student_number": str(student_no).strip()})
        course  = current_app.db.courses.find_one({"course_code": str(course_code).strip().upper()})

        if not student:
            errors.append(f"Satır {line_no}: '{student_no}' numaralı öğrenci bulunamadı.")
            error_count += 1
            continue
        if not course:
            errors.append(f"Satır {line_no}: '{course_code}' kodlu ders bulunamadı.")
            error_count += 1
            continue

        data = {
            "student_id":     str(student["_id"]),
            "course_id":      str(course["_id"]),
            "midterm_score":  midterm,
            "project_score":  project,
            "final_score":    final,
        }
        result = service.upsert_grade(data)
        if result.get("success"):
            success_count += 1
        else:
            error_count += 1
            errors.append(f"Satır {line_no}: {result.get('errors','Kaydedilemedi')}")

    return success_response(
        {"success_count": success_count, "error_count": error_count, "errors": errors[:10]},
        f"{success_count} not kaydedildi, {error_count} hata.",
        200,
    )


@grades_bp.route("/student/<student_id>/failing", methods=["GET"])
@login_required
def failing_courses(student_id: str):
    result = get_service().get_failing_courses(student_id)
    return success_response(result)


@grades_bp.route("/export/excel", methods=["GET"])
@teacher_required
def export_excel():
    """Notları Excel dosyası olarak indir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    course_id = request.args.get("course_id")
    class_name = request.args.get("class_name", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notlar"

    # Başlık stili
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="4361EE")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Öğrenci No", "Ad Soyad", "Sınıf", "Ders", "Ders Kodu", "Vize", "Proje", "Final", "Ortalama", "Harf Notu"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Veri çek
    query = {}
    if course_id:
        query["course_id"] = ObjectId(course_id)

    pipeline = [
        {"$match": query},
        {"$lookup": {"from": "students", "localField": "student_id", "foreignField": "_id", "as": "student"}},
        {"$lookup": {"from": "courses",  "localField": "course_id",  "foreignField": "_id", "as": "course"}},
        {"$unwind": {"path": "$student", "preserveNullAndEmptyArrays": True}},
        {"$unwind": {"path": "$course",  "preserveNullAndEmptyArrays": True}},
        {"$sort": {"student.full_name": 1}},
    ]
    if class_name:
        pipeline.insert(1, {"$match": {"student.class_name": class_name}})

    grades = list(current_app.db.grades.aggregate(pipeline))

    pass_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")

    for row_idx, g in enumerate(grades, 2):
        student = g.get("student", {})
        course  = g.get("course", {})
        midterm = g.get("midterm_score") or 0
        project = g.get("project_score") or 0
        final   = g.get("final_score") or 0
        avg     = round(midterm * 0.3 + (project * 0.1 if project else 0) + final * (0.6 if not project else 0.6), 1)

        if avg >= 85:   letter = "AA"
        elif avg >= 75: letter = "BA"
        elif avg >= 65: letter = "BB"
        elif avg >= 60: letter = "CB"
        elif avg >= 50: letter = "CC"
        elif avg >= 45: letter = "DC"
        elif avg >= 40: letter = "DD"
        else:           letter = "FF"

        row_fill = pass_fill if avg >= 45 else fail_fill
        row_data = [
            student.get("student_number", ""),
            student.get("full_name", ""),
            student.get("class_name", ""),
            course.get("course_name", ""),
            course.get("course_code", ""),
            midterm, project, final, avg, letter,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            if col >= 6:
                cell.fill = row_fill

    # Sütun genişlikleri
    col_widths = [14, 24, 16, 28, 12, 8, 8, 8, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"notlar_{class_name or 'tum'}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
